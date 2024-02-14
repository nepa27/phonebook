import os

import openpyxl

PHONEBOOK: str = 'phonebook.xlsx'
COUNTS_COLUMNS: int = 7
OFFSET_FOR_VIEW: int = 6
OFFSET_FOR_PHONE: int = 5


def check_or_create_excel_file() -> None:
    """Проверяет наличие файла Excel или создает новый."""
    if not os.path.exists(PHONEBOOK):
        book = openpyxl.Workbook()
        ws = book.active
        ws.title = 'Phonebook'
        ws['A1'] = 'Фамилия'
        ws['B1'] = 'Имя'
        ws['C1'] = 'Отчество'
        ws['D1'] = 'Название организации'
        ws['E1'] = 'Телефон рабочий'
        ws['F1'] = 'Телефон личный (сотовый)'
        book.save(PHONEBOOK)
        book.close()


def return_phonebook() -> tuple[list, any]:
    """Возвращает данные из телефонной книги."""
    values: list = []
    book = openpyxl.load_workbook(PHONEBOOK)
    sheet = book.active
    rows = sheet.max_row
    for row in range(2, sheet.max_row + 1):
        for col in range(1, COUNTS_COLUMNS):
            values.append(sheet.cell(row, col).value)
    book.close()

    return values, rows


def all_notes(data: list) -> None:
    """Выводит поочередно данные из телефонной книги."""
    values = data
    view_one_note(data[:OFFSET_FOR_VIEW])
    while True:
        next_note = input(
            'Нажмите 1, чтобы показать следующий контакт\n'
            'Нажмите 2, чтобы выйти в меню\n')
        if next_note == '1':
            try:
                values = values[OFFSET_FOR_VIEW:]
                view_one_note(values)
            except ValueError:
                print('-----------------------------\n'
                      'Телефонная книга закончилась!\n'
                      '-----------------------------\n')
                main()
        elif next_note == '2':
            main()
        else:
            error_command()


def view_one_note(data: list, offset: int = 0) -> bool:
    """Выводит одну запись из телефонной книги."""
    (second_name,
     name,
     surname,
     organization,
     phone,
     mob_phone) = data[offset: OFFSET_FOR_VIEW + offset]
    print(f'Фамилия: {second_name}\n'
          f'Имя: {name}\n'
          f'Отчество: {surname}\n'
          f'Название организации: {organization}\n'
          f'Телефон рабочий: {phone}\n'
          f'Телефон личный: {mob_phone}\n')

    return True


def add_note(new_data: str or list) -> any:
    """Добавляет новую запись в телефонную книгу."""
    if type(new_data) is str:
        data = new_data.split(',')
        if len(data) != 6:
            print(f'Неправильный формат данных! {data}')
            return data
    else:
        data = new_data
    phonebook = return_phonebook()[0]
    if data[0] in phonebook or data[-1]:
        print(f'\nФамилия или мобильный номер ({data[0]})'
              f' уже есть в телефонной книге!\n')
        return
    book = openpyxl.load_workbook(PHONEBOOK)
    ws = book.active
    ws.append(data)
    book.save(PHONEBOOK)
    book.close()

    return data


def change_note(second_name: str) -> None:
    """Изменяет запись в телефонной книге."""
    values = search_note(second_name)
    try:
        right_index = values.index(second_name)
    except ValueError:
        return
    new_data = input('Введите через запятую:\n'
                     'Фамилию, имя, отчество, название организации, '
                     'рабочий телефон, личный\n')
    new_data = new_data.split(',')
    data = return_phonebook()
    for index, element in enumerate(data[0][right_index:]):
        for ind, el in enumerate(new_data):
            if index == ind + right_index:
                data[0].pop(index)
                data[0].insert(index, el)
    delete_data()
    offset: int = 0
    rows: int = 0
    while rows != data[1] - 1:
        add_note(data[0][offset:offset + 6])
        offset += 6
        rows += 1

    return


def search_index(data: list, value: str, offset: int = 0) -> None:
    """Ищет индекс значения в списке."""
    try:
        right_index = data.index(value)
        print('-----------------------------\n')
        result = view_one_note(data, right_index - offset)
        if result:
            print('Запись найдена!\n')
    except ValueError:
        print('\nЗапись не найдена!\n')


def search_note(second_mane: str, mob_phone: str = '') -> list:
    """Ищет запись в телефонной книге."""
    data = return_phonebook()
    if second_mane:
        search_index(data[0], second_mane)
    elif mob_phone:
        search_index(data[0], mob_phone, OFFSET_FOR_PHONE)
    else:
        print('\nЗапись не найдена!\n')

    return data[0]


def delete_data() -> None:
    """Очищает телефонную книгу. """
    book = openpyxl.load_workbook(PHONEBOOK)
    sheet = book.active
    range_ = range(2, sheet.max_row + 1)
    for row in reversed(list(range_)):
        sheet.delete_rows(row)
    book.save(PHONEBOOK)
    book.close()


def error_command() -> None:
    """Отображает оповещение об ошибке и вызывает main()."""
    print('Неверная команда!')
    main()


def main() -> None:
    """Главная функция."""
    check_or_create_excel_file()
    choose = input('-----------------------------\n'
                   '      Телефонная книга.\n'
                   '-----------------------------\n'
                   'Нажмите 1, чтобы посмотреть содержимое\n'
                   'Нажмите 2, чтобы добавить запись\n'
                   'Нажмите 3, чтобы изменить запись\n'
                   'Нажмите 4, чтобы найти запись\n'
                   'Нажмите 5, чтобы выйти\n'
                   '-----------------------------\n')
    if choose == '1':
        all_notes(return_phonebook()[0])
    elif choose == '2':
        new_note = input('Введите ЧЕРЕЗ ЗАПЯТУЮ:\n'
                         'Фамилию, имя, отчество, название организации, '
                         'рабочий телефон, личный\n')
        add_note(new_note)
        print('Запись успешно добавлена!')
        main()
    elif choose == '3':
        second_name = input('Введите фамилию человека, '
                            'чью запись хотите изменить: ')
        change_note(second_name)
        main()
    elif choose == '4':
        print('Введите фамилию или номер телефона '
              '(заполните только то поле,\nпо которому '
              'хотите произвести поиск, '
              'другое поле оставьте пустым):')
        second_mane = input('Фамилия: ')
        mob_phone = input('Мобильный номер: ')
        search_note(second_mane, mob_phone)
        main()
    elif choose == '5':
        exit()
    else:
        error_command()


if __name__ == "__main__":
    main()
